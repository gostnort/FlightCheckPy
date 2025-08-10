#!/usr/bin/env python3
"""
Excel处理页面 - 导入Excel文件并根据TKNE和CKIN CCRD生成输出文件
"""

import streamlit as st
import pandas as pd
import sqlite3
import os
import re
from typing import List, Dict, Tuple, Optional
from ui.common import apply_global_settings, get_current_database
from scripts.hbpr_info_processor import HbprDatabase




def show_excel_processor():
    """显示Excel处理页面"""
    apply_global_settings()
    
    # 检查数据库状态
    selected_db_file = get_current_database()
    if not selected_db_file:
        st.error("❌ 未选择数据库!")
        st.info("💡 请从侧边栏选择数据库或先创建数据库。")
        return
    
    db = HbprDatabase(selected_db_file)
    st.success(f"✅ 数据库已连接: {os.path.basename(selected_db_file)}")
    
    # 文件上传区域
    st.subheader("📁 上传Excel文件")
    uploaded_file = st.file_uploader(
        "选择要处理的Excel文件 (例如: sample_in_25JUL.xlsx)",
        type=['xlsx', 'xls'],
        help="上传包含TKNE数据的Excel文件进行处理"
    )
    
    if uploaded_file is not None:
        try:
            # 正确读取Excel文件，跳过标题行，支持XLS和XLSX格式
            file_ext = uploaded_file.name.lower().split('.')[-1]
            
            if file_ext == 'xls':
                # 对于XLS格式，明确指定引擎
                try:
                    df_input = pd.read_excel(uploaded_file, skiprows=1, engine='xlrd')
                    st.info("📊 检测到XLS格式文件，使用xlrd引擎读取")
                except ImportError:
                    st.error("❌ 缺少xlrd包，无法读取XLS文件。请安装：pip install xlrd")
                    return
                except Exception as e:
                    st.error(f"❌ 读取XLS文件失败: {str(e)}")
                    return
            else:
                # 对于XLSX格式，使用默认引擎
                try:
                    df_input = pd.read_excel(uploaded_file, skiprows=1, engine='openpyxl')
                    st.info("📊 检测到XLSX格式文件，使用openpyxl引擎读取")
                except Exception as e:
                    st.error(f"❌ 读取XLSX文件失败: {str(e)}")
                    return
            
            st.success(f"✅ 成功读取文件: {uploaded_file.name}")
            
            # 检查必要的列是否存在
            required_columns = ['EMD', '关联ET', '旅客姓名', '航班号', '航程', '操作', '实收金额', '工作号', '操作时间', '产品类型']
            missing_columns = [col for col in required_columns if col not in df_input.columns]
            
            if missing_columns:
                st.error(f"❌ 缺少必要的列: {missing_columns}")
                st.info("💡 请确保Excel文件是EMD销售日报格式，包含所有必要的列")
                return
            
            # 处理按钮
            if st.button("🚀 开始处理", type="primary", use_container_width=True):
                with st.spinner("正在处理Excel文件..."):
                    result_df, unprocessed_records = process_excel_file(df_input, db)
                    
                if result_df is not None:
                    # 显示处理结果
                    st.subheader("✅ 处理结果")
                    st.dataframe(result_df, use_container_width=True)
                    
                    # 显示未处理的记录
                    if unprocessed_records:
                        st.subheader("⚠️ 未处理的CKIN CCRD记录")
                        for record in unprocessed_records:
                            st.warning(f"乘客: {record['name']}, TKNE: {record['tkne']}, CKIN CCRD: {record['ckin_ccrd']}")
                    
                    # 生成输出文件，保存到合适的位置
                    from datetime import datetime
                    
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"processed_output_{uploaded_file.name.replace('.xlsx', '')}_{timestamp}.xlsx"
                    
                    # 确定保存路径
                    output_file = get_output_file_path(filename)
                    generate_output_excel(result_df, unprocessed_records, output_file)
                    
                    # 显示文件保存位置和提供下载链接
                    st.subheader("📥 文件已生成")
                    st.success(f"✅ 文件已保存到: {output_file}")
                    
                    # 提供下载链接作为备用方案
                    with open(output_file, 'rb') as f:
                        st.download_button(
                            label="📥 备用下载 (如果文件位置无法访问)",
                            data=f.read(),
                            file_name=f"processed_{uploaded_file.name}",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )  
        except Exception as e:
            st.error(f"❌ 处理文件时发生错误: {str(e)}")
            st.info("💡 请检查Excel文件格式是否正确")




def process_excel_file(df_input: pd.DataFrame, db: HbprDatabase) -> Tuple[Optional[pd.DataFrame], List[Dict]]:
    """
    处理Excel文件，根据TKNE查找数据库记录并生成输出
    
    Args:
        df_input: 输入的DataFrame
        db: 数据库连接对象
    
    Returns:
        Tuple[处理后的DataFrame, 未处理的记录列表]
    """
    try:
        # 初始化输出DataFrame
        output_data = []
        unprocessed_records = []
        
        # 获取所有包含CKIN CCRD的HBNB记录
        hbnb_list = get_all_ckin_ccrd_hbnb(db)
        st.info(f"📊 数据库中找到 {len(hbnb_list)} 个包含CKIN CCRD的记录")
        
        # 处理每一行输入数据
        for index, row in df_input.iterrows():
            try:
                # 从'关联ET'列获取TKNE号码
                tkne = str(row.get('关联ET', '')).strip()
                if not tkne or tkne == 'nan' or tkne == '':
                    continue
                
                # 在数据库中查找对应的TKNE记录
                hbnb_records = find_records_by_tkne(db, tkne)
                
                # 创建基础输出行
                output_row = create_base_output_row(row)
                
                # 处理CKIN CCRD信息
                for hbnb_record in hbnb_records:
                    if has_ckin_ccrd(hbnb_record):
                        ckin_data = parse_ckin_ccrd(hbnb_record['ckin_msg'])
                        if ckin_data['success']:
                            # 成功解析CKIN CCRD，添加到输出行
                            output_row.update(ckin_data['data'])
                            
                            # 从hbnb_list中移除已处理的记录
                            hbnb_list = [h for h in hbnb_list if h['hbnb_number'] != hbnb_record['hbnb_number']]
                            break
                        else:
                            # 解析失败，添加到未处理记录
                            unprocessed_records.append({
                                'name': hbnb_record.get('name', '未知'),
                                'tkne': tkne,
                                'ckin_ccrd': hbnb_record['ckin_msg']
                            })
                
                output_data.append(output_row)
                
            except Exception as e:
                st.warning(f"⚠️ 处理第 {index+1} 行时出错: {str(e)}")
                continue
        
        # 将剩余未处理的CKIN CCRD添加到未处理记录
        for remaining_hbnb in hbnb_list:
            unprocessed_records.append({
                'name': remaining_hbnb.get('name', '未知'),
                'tkne': remaining_hbnb.get('tkne', '未知'),
                'ckin_ccrd': remaining_hbnb['ckin_msg']
            })
        
        # 创建输出DataFrame
        if output_data:
            result_df = pd.DataFrame(output_data)
            return result_df, unprocessed_records
        else:
            st.warning("⚠️ 没有生成任何输出数据")
            return None, unprocessed_records
            
    except Exception as e:
        st.error(f"❌ 处理过程中发生错误: {str(e)}")
        return None, []




def get_all_ckin_ccrd_hbnb(db: HbprDatabase) -> List[Dict]:
    """获取所有包含CKIN CCRD的HBNB记录"""
    try:
        conn = sqlite3.connect(db.db_file)
        cursor = conn.cursor()
        
        query = """
        SELECT hbnb_number, name, tkne, ckin_msg 
        FROM hbpr_full_records 
        WHERE ckin_msg IS NOT NULL 
        AND ckin_msg LIKE '%CKIN CCRD%'
        AND ckin_msg != ''
        """
        
        cursor.execute(query)
        records = cursor.fetchall()
        conn.close()
        
        return [
            {
                'hbnb_number': record[0],
                'name': record[1] or '未知',
                'tkne': record[2] or '',
                'ckin_msg': record[3] or ''
            }
            for record in records
        ]
        
    except Exception as e:
        st.error(f"❌ 查询CKIN CCRD记录时出错: {str(e)}")
        return []




def find_records_by_tkne(db: HbprDatabase, tkne: str) -> List[Dict]:
    """根据TKNE查找数据库记录"""
    try:
        conn = sqlite3.connect(db.db_file)
        cursor = conn.cursor()
        
        # 清理TKNE格式：移除.0后缀，准备多种匹配格式
        clean_tkne = str(tkne).replace('.0', '') if tkne else ''
        
        query = """
        SELECT hbnb_number, name, tkne, ckin_msg 
        FROM hbpr_full_records 
        WHERE tkne LIKE ? OR tkne LIKE ? OR tkne = ?
        """
        
        # 多种匹配模式：包含/1后缀，包含/2后缀，精确匹配
        patterns = [
            f'{clean_tkne}/1',
            f'{clean_tkne}/2', 
            clean_tkne
        ]
        
        cursor.execute(query, patterns)
        records = cursor.fetchall()
        conn.close()
        
        return [
            {
                'hbnb_number': record[0],
                'name': record[1] or '未知',
                'tkne': record[2] or '',
                'ckin_msg': record[3] or ''
            }
            for record in records
        ]
        
    except Exception as e:
        st.error(f"❌ 查询TKNE记录时出错: {str(e)}")
        return []




def has_ckin_ccrd(record: Dict) -> bool:
    """检查记录是否包含CKIN CCRD"""
    ckin_msg = record.get('ckin_msg', '')
    return 'CKIN CCRD' in ckin_msg




def create_base_output_row(input_row: pd.Series) -> Dict:
    """创建基础输出行，包含列映射和固定值"""
    # 根据实际EMD销售日报的列映射关系
    output_row = {}
    
    try:
        # 根据request.md的映射关系，使用实际的列名
        # 原mapping: B->A, C->B, F->C, J->E, R->F, S->G, T->H, E->J, K->K
        
        output_row['A'] = str(input_row.get('EMD', ''))                    # EMD -> A列 (原B列)
        output_row['B'] = str(input_row.get('旅客姓名', ''))               # 旅客姓名 -> B列 (原C列)
        output_row['C'] = str(input_row.get('航程', ''))                   # 航程 -> C列 (原F列)
        # 操作 -> E列的翻译处理 (原J列)
        operation_value = str(input_row.get('操作', ''))
        output_row['E'] = translate_operation_to_english(operation_value)
        output_row['F'] = str(input_row.get('工作号', ''))                 # 工作号 -> F列 (原R列)
        output_row['G'] = str(input_row.get('操作时间', ''))               # 操作时间 -> G列 (原S列)
        output_row['J'] = str(input_row.get('航班号', ''))                 # 航班号 -> J列 (原E列)
        output_row['K'] = str(input_row.get('实收金额', ''))               # 实收金额 -> K列 (原K列)
        
        # 产品类型 -> H列的翻译处理 (原T列)
        product_type = str(input_row.get('产品类型', ''))
        output_row['H'] = translate_column_t_to_h(product_type)
        
        # 固定值
        output_row['D'] = "1"
        output_row['I'] = "International"
        
        # 初始化CKIN CCRD相关列
        output_row['L'] = ""  # CASH类型
        output_row['M'] = ""  # AX类型的ITEM2
        output_row['N'] = ""  # 其他类型的ITEM2  
        output_row['O'] = ""  # 4位数字
        output_row['P'] = ""  # ITEM3及后续内容
        
    except Exception as e:
        st.warning(f"⚠️ 创建输出行时出错: {str(e)}")
        # 如果出错，创建空的基础行
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
            output_row[col] = ""
        output_row['D'] = "1"
        output_row['I'] = "International"
    
    return output_row




def translate_column_t_to_h(value: str) -> str:
    """将T列内容翻译为H列内容"""
    if not value or str(value).strip() == '' or str(value) == 'nan':
        return ""
    
    value_str = str(value).strip()
    
    # 翻译映射
    translation_map = {
        '逾重PC': 'EXPC',
        '选座': 'SEAT', 
        '升舱': 'UPG'
    }
    
    return translation_map.get(value_str, value_str)




def translate_operation_to_english(value: str) -> str:
    """将操作列内容翻译为英文"""
    if not value or str(value).strip() == '' or str(value) == 'nan':
        return ""
    
    value_str = str(value).strip()
    
    # 操作翻译映射
    operation_translation = {
        '出票': 'Issue'
    }
    
    return operation_translation.get(value_str, value_str)




def parse_ckin_ccrd(ckin_msg: str) -> Dict:
    """
    解析CKIN CCRD信息
    
    Returns:
        Dict: {
            'success': bool,
            'data': Dict  # 包含L, M, N, O, P列的数据
        }
    """
    try:
        if not ckin_msg or 'CKIN CCRD' not in ckin_msg:
            return {'success': False, 'data': {}}
        
        # 查找所有CKIN CCRD部分（可能有多个）
        ckin_pattern = r'CKIN CCRD\s+([^;]+)'
        matches = re.findall(ckin_pattern, ckin_msg, re.IGNORECASE)
        
        if not matches:
            return {'success': False, 'data': {}}
        
        data = {'L': '', 'M': '', 'N': '', 'O': '', 'P': ''}
        
        # 处理每个CKIN CCRD条目
        for ccrd_content in matches:
            ccrd_content = ccrd_content.strip()
            parts = ccrd_content.split()
            
            if len(parts) < 1:
                continue
                
            item1 = parts[0]
            item2 = parts[1] if len(parts) > 1 else ''
            item3_and_beyond = ' '.join(parts[2:]) if len(parts) > 2 else ''
            
            # 处理CASH类型
            if item1.upper() == "CASH":
                data['L'] = item2
                return {'success': True, 'data': data}
                
            # 处理2字母+4数字格式
            elif re.match(r'^[A-Z]{2}\d{4}$', item1):
                letters = item1[:2]
                digits = item1[2:]
                
                data['O'] = digits  # 4位数字写入O列
                
                if letters.upper() == 'AX':
                    # AX类型：ITEM2写入M列
                    data['M'] = item2
                else:
                    # 其他类型（如VI）：ITEM2写入N列
                    data['N'] = item2
                
                # ITEM3及后续写入P列
                data['P'] = item3_and_beyond
                return {'success': True, 'data': data}
        
        # 如果没有找到符合格式的CCRD，但有CCRD内容，返回失败以便记录
        return {'success': False, 'data': {}}
        
    except Exception:
        return {'success': False, 'data': {}}




def extract_ckin_ccrd_content(ckin_msg: str) -> str:
    """提取CKIN CCRD后面的内容直到分号"""
    try:
        if not ckin_msg or 'CKIN CCRD' not in ckin_msg:
            return ckin_msg
        
        # 查找CKIN CCRD部分并提取到分号结束
        import re
        pattern = r'CKIN CCRD\s+([^;]+)'
        match = re.search(pattern, ckin_msg)
        
        if match:
            return match.group(1).strip()
        else:
            return ckin_msg
            
    except Exception:
        return ckin_msg




def number_to_english(amount: float) -> str:
    """将数字转换为英文金额表示"""
    try:
        # 简单的数字转英文实现
        ones = ['', 'ONE', 'TWO', 'THREE', 'FOUR', 'FIVE', 'SIX', 'SEVEN', 'EIGHT', 'NINE']
        teens = ['TEN', 'ELEVEN', 'TWELVE', 'THIRTEEN', 'FOURTEEN', 'FIFTEEN', 'SIXTEEN', 'SEVENTEEN', 'EIGHTEEN', 'NINETEEN']
        tens = ['', '', 'TWENTY', 'THIRTY', 'FORTY', 'FIFTY', 'SIXTY', 'SEVENTY', 'EIGHTY', 'NINETY']
        
        def convert_hundreds(n):
            result = ''
            if n >= 100:
                result += ones[n // 100] + ' HUNDRED '
                n %= 100
            if n >= 20:
                result += tens[n // 10] + ' '
                n %= 10
            elif n >= 10:
                result += teens[n - 10] + ' '
                n = 0
            if n > 0:
                result += ones[n] + ' '
            return result.strip()
        
        # 分离整数和小数部分
        dollars = int(amount)
        cents = int(round((amount - dollars) * 100))
        
        result = ''
        
        if dollars == 0:
            result = 'ZERO DOLLARS'
        else:
            # 处理千位
            if dollars >= 1000:
                thousands = dollars // 1000
                result += convert_hundreds(thousands) + ' THOUSAND '
                dollars %= 1000
            
            # 处理百位
            if dollars > 0:
                result += convert_hundreds(dollars)
            
            # 处理单复数
            if int(amount) == 1:
                result += ' DOLLAR'
            else:
                result += ' DOLLARS'
        
        # 处理分
        if cents > 0:
            if cents == 1:
                result += ' AND ' + convert_hundreds(cents) + ' CENT'
            else:
                result += ' AND ' + convert_hundreds(cents) + ' CENTS'
        else:
            result += ' AND NO CENTS'
        
        result += ' EXACTLY'
        
        return result.strip()
        
    except Exception:
        return f"${amount:.2f}"




def get_output_file_path(filename: str) -> str:
    """确定输出文件的保存路径"""
    
    # 首先尝试用户的Downloads文件夹
    downloads_path = os.path.expanduser("~/Downloads")
    
    if os.path.exists(downloads_path) and os.access(downloads_path, os.W_OK):
        output_path = os.path.join(downloads_path, filename)
        st.info(f"📁 文件将保存到: Downloads/{filename}")
        return output_path
    
    # 如果Downloads不存在或无法访问，尝试创建C:\temp
    try:
        temp_dir = "C:\\temp"
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)
            st.info(f"📁 创建临时目录: {temp_dir}")
        
        output_path = os.path.join(temp_dir, filename)
        st.info(f"📁 文件将保存到: {temp_dir}\\{filename}")
        return output_path
        
    except Exception as e:
        # 最后的备用方案：当前工作目录
        st.warning(f"⚠️ 无法访问Downloads或创建C:\\temp，使用当前目录: {str(e)}")
        return filename




def generate_output_excel(result_df: pd.DataFrame, unprocessed_records: List[Dict], output_file: str) -> str:
    """基于Out_format.xlsx模板生成输出Excel文件"""
    try:
        from openpyxl import load_workbook
        
        template_file = "resources/Out_format.xlsx"
        
        # 检查模板文件是否存在
        if not os.path.exists(template_file):
            st.error(f"❌ 模板文件不存在: {template_file}")
            return None
        
        # 加载模板文件
        wb = load_workbook(template_file)
        
        # 使用EMD工作表
        if 'EMD' not in wb.sheetnames:
            st.error("❌ 模板文件中没有EMD工作表")
            return None
        
        ws_emd = wb['EMD']
        
        # 从A8开始写入数据
        start_row = 8
        headers = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']
        
        # 写入数据到EMD工作表
        for data_idx, (_, row) in enumerate(result_df.iterrows()):
            row_idx = start_row + data_idx
            for col_idx, header in enumerate(headers, 1):
                value = row.get(header, '')
                
                # 特殊处理数字列 (K=11, L=12, M=13, N=14)，但不包括O=15列
                if col_idx in [11, 12, 13, 14]:  # K, L, M, N列需要数字格式
                    try:
                        # 如果是数字，转换为float；如果是空或非数字，设为0
                        if value and str(value).strip() and str(value) != 'nan':
                            numeric_value = float(str(value).strip())
                        else:
                            numeric_value = 0
                        ws_emd.cell(row=row_idx, column=col_idx, value=numeric_value)
                    except (ValueError, TypeError):
                        ws_emd.cell(row=row_idx, column=col_idx, value=0)
                elif col_idx == 15:  # O列：空值保持为空，不写0
                    if value and str(value).strip() and str(value) != 'nan':
                        ws_emd.cell(row=row_idx, column=col_idx, value=str(value).strip())
                    else:
                        ws_emd.cell(row=row_idx, column=col_idx, value='')
                else:
                    # 其他列保持文本格式
                    ws_emd.cell(row=row_idx, column=col_idx, value=str(value) if value else '')
        
        # 处理SUM工作表的未处理记录和其他信息
        if 'SUM' not in wb.sheetnames:
            ws_sum = wb.create_sheet('SUM')
        else:
            ws_sum = wb['SUM']
        
        # 写入航班号数字部分到K4
        if len(result_df) > 0:
            # 从第一行数据获取航班号
            flight_number = result_df.iloc[0].get('J', '')  # J列是航班号
            if flight_number and isinstance(flight_number, str):
                # 提取数字部分，如CA984 -> 984
                import re
                flight_digits = re.findall(r'\d+', flight_number)
                if flight_digits:
                    ws_sum.cell(row=4, column=11, value=flight_digits[0])  # K4
        
        # 写入航班日期到C14 (假设使用当前日期，可以根据实际需求调整)
        from datetime import datetime
        flight_date = datetime.now().strftime('%Y-%m-%d')
        ws_sum.cell(row=14, column=3, value=flight_date)  # C14
        
        # 从C15开始添加未处理记录（只写CKIN CCRD内容，不显示"CKIN CCRD"字样）
        if unprocessed_records:
            row_idx = 15
            for record in unprocessed_records:
                # 提取CKIN CCRD后面的内容
                ckin_ccrd_content = extract_ckin_ccrd_content(record['ckin_ccrd'])
                # 不显示"CKIN CCRD:"，直接显示内容
                error_info = f"乘客: {record['name']}, TKNE: {record['tkne']}, {ckin_ccrd_content}"
                ws_sum.cell(row=row_idx, column=3, value=error_info)  # C列
                row_idx += 1
        
        # 处理Receipt工作表
        if 'RECEIPT' in wb.sheetnames:
            ws_receipt = wb['RECEIPT']
            
            # 计算现金总额(L列是第12列，现金数据)
            cash_total = 0.0
            for _, row in result_df.iterrows():
                l_value = row.get('L', '')
                if l_value and str(l_value).strip() and str(l_value).strip() != 'nan':
                    try:
                        cash_amount = float(str(l_value).strip())
                        cash_total += cash_amount
                    except ValueError:
                        continue
            
            if cash_total > 0:
                # 转换数字为英文
                english_amount = number_to_english(cash_total)
                # 写入Receipt表的C8位置
                ws_receipt.cell(row=8, column=3, value=english_amount)  # C8位置
                st.info(f"💰 已将现金总额 ${cash_total:.2f} 转换为英文写入Receipt表C8")
        
        # 保存到新文件
        wb.save(output_file)
        wb.close()
        
        return output_file
        
    except Exception as e:
        st.error(f"❌ 生成输出文件时出错: {str(e)}")
        import traceback
        st.error(f"详细错误: {traceback.format_exc()}")
        return None
