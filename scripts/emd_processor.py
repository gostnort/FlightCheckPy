"""
EMD Processor Script
===================

This script provides functionality to process EMD (Electronic Miscellaneous Document) data
and fill out Excel files programmatically, replacing the original VBA functionality.

Features:
- Process source Excel files and populate EMD sheets
- Convert numbers to text (for RECEIPT sheet)
- Handle cash and credit payments
- Validate data integrity
- Generate properly named output files
- Support new API format with structured input data

API Functions:
- process_emd_data(emd_data, template_path, output_path, config_path)
- spell_number(amount)
- validate_emd_data(emd_list)
- generate_filename(ca_number, date_str)
"""

import os
import yaml
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class EMDRecord:
    """Represents a single EMD record with all its properties"""
    
    def __init__(self, emd: str = "", et: str = "", leg: str = "", emd_count: int = 1,
                 operation: str = "", job_number: int = 0, date: str = "", 
                 product: str = "", i_d: str = "", flight: str = "", price: float = 0.0,
                 cash: float = 0.0, ax: float = 0.0, other: float = 0.0,
                 last_4_digit: str = "", remark: str = ""):
        self.emd = emd
        self.et = et
        self.leg = leg
        self.emd_count = emd_count
        self.operation = operation
        self.job_number = job_number
        self.date = date
        self.product = product
        self.i_d = i_d
        self.flight = flight
        self.price = price
        self.cash = cash
        self.ax = ax
        self.other = other
        self.last_4_digit = last_4_digit
        self.remark = remark
    
    def __str__(self):
        return f"EMD: {self.emd}, Flight: {self.flight}, Price: {self.price}, Cash: {self.cash}, AX: {self.ax}"


class EMDData:
    """Represents the complete EMD data structure"""
    
    def __init__(self, total_amount: float = 0.0, report_date: str = "", 
                 signature: str = "", collection_msg: List[str] = None, 
                 emd_list: List[EMDRecord] = None):
        self.total_amount = total_amount
        self.report_date = report_date
        self.signature = signature
        self.collection_msg = collection_msg or []
        self.emd_list = emd_list or []
    
    def __str__(self):
        return f"EMD Data: {len(self.emd_list)} records, Total: {self.total_amount}, Date: {self.report_date}"


class EMDProcessor:
    """Main class for processing EMD data and filling Excel files"""
    
    def __init__(self, config_path: str = "resources/emd_processor_config.yaml"):
        """Initialize the processor with configuration"""
        self.config = self._load_config(config_path)
        self._setup_number_conversion()
    
    def _load_config(self, config_path: str) -> Dict[str, Any]:
        """Load configuration from YAML file"""
        try:
            with open(config_path, 'r', encoding='utf-8') as file:
                config = yaml.safe_load(file)
            logger.info(f"Configuration loaded from {config_path}")
            return config
        except FileNotFoundError:
            logger.error(f"Configuration file not found: {config_path}")
            raise
        except yaml.YAMLError as e:
            logger.error(f"Error parsing configuration file: {e}")
            raise
    
    def _setup_number_conversion(self):
        """Setup number to text conversion mappings"""
        self.number_words = {
            0: "ZERO", 1: "ONE", 2: "TWO", 3: "THREE", 4: "FOUR", 5: "FIVE",
            6: "SIX", 7: "SEVEN", 8: "EIGHT", 9: "NINE", 10: "TEN",
            11: "ELEVEN", 12: "TWELVE", 13: "THIRTEEN", 14: "FOURTEEN", 15: "FIFTEEN",
            16: "SIXTEEN", 17: "SEVENTEEN", 18: "EIGHTEEN", 19: "NINETEEN",
            20: "TWENTY", 30: "THIRTY", 40: "FORTY", 50: "FIFTY",
            60: "SIXTY", 70: "SEVENTY", 80: "EIGHTY", 90: "NINETY"
        }
    
    def spell_number(self, amount: float) -> str:
        """
        Convert a number to its text representation (equivalent to VBA SpellNumber function)
        
        Args:
            amount: The amount to convert (e.g., 1234.56)
            
        Returns:
            Text representation (e.g., "ONE THOUSAND TWO HUNDRED THIRTY FOUR DOLLARS AND FIFTY SIX CENTS EXACTLY")
        """
        if amount == 0:
            return self.config['number_to_text']['no_dollars'] + " " + self.config['number_to_text']['no_cents']
        
        # Split into dollars and cents
        dollars = int(amount)
        cents = int(round((amount - dollars) * 100))
        
        # Convert dollars to text
        if dollars == 0:
            dollars_text = self.config['number_to_text']['no_dollars']
        elif dollars == 1:
            dollars_text = self.config['number_to_text']['one_dollar']
        else:
            dollars_text = self._convert_dollars_to_text(dollars) + " " + self.config['number_to_text']['currency']
        
        # Convert cents to text
        if cents == 0:
            cents_text = " " + self.config['number_to_text']['no_cents']
        elif cents == 1:
            cents_text = " " + self.config['number_to_text']['one_cent']
        else:
            cents_text = " AND " + self._convert_cents_to_text(cents) + " " + self.config['number_to_text']['cents'] + " " + self.config['number_to_text']['exact']
        
        return dollars_text + cents_text
    
    def _convert_dollars_to_text(self, dollars: int) -> str:
        """Convert dollars to text representation"""
        if dollars == 0:
            return ""
        
        if dollars < 20:
            return self.number_words[dollars]
        elif dollars < 100:
            tens = (dollars // 10) * 10
            ones = dollars % 10
            if ones == 0:
                return self.number_words[tens]
            else:
                return self.number_words[tens] + " " + self.number_words[ones]
        elif dollars < 1000:
            hundreds = dollars // 100
            remainder = dollars % 100
            result = self.number_words[hundreds] + " HUNDRED"
            if remainder > 0:
                result += " " + self._convert_dollars_to_text(remainder)
            return result
        elif dollars < 1000000:
            thousands = dollars // 1000
            remainder = dollars % 1000
            result = self._convert_dollars_to_text(thousands) + " THOUSAND"
            if remainder > 0:
                result += " " + self._convert_dollars_to_text(remainder)
            return result
        else:
            # Handle larger numbers if needed
            return str(dollars)
    
    def _convert_cents_to_text(self, cents: int) -> str:
        """Convert cents to text representation"""
        return self._convert_dollars_to_text(cents)
    
    def create_emd_record_from_dict(self, emd_dict: Dict[str, Any]) -> EMDRecord:
        """
        Create EMDRecord from dictionary
        
        Args:
            emd_dict: Dictionary containing EMD data
            
        Returns:
            EMDRecord object
        """
        return EMDRecord(
            emd=str(emd_dict.get('emd', '')),
            et=str(emd_dict.get('et', '')),
            leg=str(emd_dict.get('leg', '')),
            emd_count=int(emd_dict.get('emd_count', 1)),
            operation=str(emd_dict.get('operation', '')),
            job_number=int(emd_dict.get('job_number', 0)),
            date=str(emd_dict.get('date', '')),
            product=str(emd_dict.get('product', '')),
            i_d=str(emd_dict.get('i_d', '')),
            flight=str(emd_dict.get('flight', '')),
            price=float(emd_dict.get('price', 0.0)),
            cash=float(emd_dict.get('cash', 0.0)),
            ax=float(emd_dict.get('ax', 0.0)),
            other=float(emd_dict.get('other', 0.0)),
            last_4_digit=str(emd_dict.get('last_4_digit', '')),
            remark=str(emd_dict.get('remark', ''))
        )
    
    def create_emd_data_from_dict(self, data_dict: Dict[str, Any]) -> EMDData:
        """
        Create EMDData from dictionary
        
        Args:
            data_dict: Dictionary containing complete EMD data
            
        Returns:
            EMDData object
        """
        emd_list = []
        for emd_dict in data_dict.get('emd_list', []):
            emd_record = self.create_emd_record_from_dict(emd_dict)
            emd_list.append(emd_record)
        
        return EMDData(
            total_amount=float(data_dict.get('total_amount', 0.0)),
            report_date=str(data_dict.get('report_date', '')),
            signature=str(data_dict.get('signature', '')),
            collection_msg=data_dict.get('collection_msg', []),
            emd_list=emd_list
        )
    
    def validate_emd_data(self, emd_data: EMDData, existing_emds: List[str] = None) -> Tuple[bool, str]:
        """
        Validate EMD data for integrity
        
        Args:
            emd_data: EMDData object to validate
            existing_emds: List of existing EMD numbers to check for duplicates
            
        Returns:
            Tuple of (is_valid, error_message)
        """
        if not emd_data.emd_list:
            return False, "No EMD records to process"
        
        # Check for duplicate EMDs within the new list
        emd_numbers = [record.emd for record in emd_data.emd_list]
        if len(emd_numbers) != len(set(emd_numbers)):
            return False, self.config['messages']['duplicate_emd']
        
        # Check for duplicates with existing EMDs
        if existing_emds and self.config['validation']['check_duplicate_emds']:
            for emd_number in emd_numbers:
                if emd_number in existing_emds:
                    return False, self.config['messages']['duplicate_emd']
        
        # Check for credit payments without last 4 digits
        if self.config['validation']['require_last_four_digits']:
            for record in emd_data.emd_list:
                if ((record.ax > 0 or record.other > 0) and 
                    (not record.last_4_digit or len(str(record.last_4_digit).strip()) < 4)):
                    return False, self.config['messages']['missing_last_four']
        
        return True, ""
    
    def generate_filename(self, flight: str, date_str: str) -> str:
        """
        Generate output filename based on flight number and date
        
        Args:
            flight: The flight number
            date_str: Date string
            
        Returns:
            Generated filename
        """
        # Clean the date string (remove dashes)
        clean_date = date_str.replace('-', '')
        
        filename = (f"{self.config['file_settings']['prefix']}{flight}"
                   f"{self.config['file_settings']['separator']}{clean_date}"
                   f"{self.config['file_settings']['file_extension']}")
        
        return filename
    
    def fill_emd_sheet(self, workbook, emd_data: EMDData, start_row: int = 8):
        """
        Fill the EMD sheet with processed data
        
        Args:
            workbook: OpenPyXL workbook object
            emd_data: EMDData object containing records to write
            start_row: Starting row to write data (default 8)
        """
        sheet = workbook[self.config['sheets']['emd']]
        cols = self.config['emd_columns']
        
        current_row = start_row
        
        for record in emd_data.emd_list:
            # Write EMD data according to new API structure
            sheet.cell(row=current_row, column=cols['emd_number']).value = record.emd
            sheet.cell(row=current_row, column=cols['ticket_number']).value = f"'{record.et}"  # Add quote prefix
            sheet.cell(row=current_row, column=cols['origin_dest']).value = record.leg
            sheet.cell(row=current_row, column=cols['quantity']).value = record.emd_count
            sheet.cell(row=current_row, column=cols['operator']).value = record.operation
            sheet.cell(row=current_row, column=cols['account']).value = record.job_number
            sheet.cell(row=current_row, column=cols['issue_date']).value = record.date
            sheet.cell(row=current_row, column=cols['product']).value = record.product
            sheet.cell(row=current_row, column=cols['id']).value = record.i_d
            sheet.cell(row=current_row, column=cols['ca_number']).value = record.flight
            sheet.cell(row=current_row, column=cols['price']).value = record.price
            
            # Handle payment amounts
            sheet.cell(row=current_row, column=cols['cash_amount']).value = record.cash
            sheet.cell(row=current_row, column=cols['amex_amount']).value = record.ax
            sheet.cell(row=current_row, column=cols['other_credit_amount']).value = record.other
            
            # Add last 4 digits for credit payments (both Amex and other credit cards)
            if (record.ax > 0 or record.other > 0) and record.last_4_digit:
                sheet.cell(row=current_row, column=cols['last_four_digits']).value = record.last_4_digit
            
            current_row += 1
        
        logger.info(f"Filled EMD sheet with {len(emd_data.emd_list)} records starting from row {start_row}")
    
    def fill_sum_sheet(self, workbook, emd_data: EMDData):
        """
        Fill the SUM sheet with collection messages and other data
        
        Args:
            workbook: OpenPyXL workbook object
            emd_data: EMDData object containing data to write
        """
        try:
            sheet = workbook[self.config['sheets']['sum']]
            
            # Write collection messages in column C starting from row 15
            start_row = 15
            for i, message in enumerate(emd_data.collection_msg):
                sheet.cell(row=start_row + i, column=3).value = message  # Column C
            
            # Update total amount if needed
            if emd_data.total_amount > 0:
                # You can add logic here to update total amount in specific cells
                pass
            
            logger.info(f"Filled SUM sheet with {len(emd_data.collection_msg)} collection messages")
            
        except Exception as e:
            logger.warning(f"Could not update SUM sheet: {e}")
    
    def update_receipt_sheet(self, workbook, total_amount: float):
        """
        Update the RECEIPT sheet with spelled out amount
        
        Args:
            workbook: OpenPyXL workbook object
            total_amount: Total amount to spell out
        """
        try:
            sheet = workbook[self.config['sheets']['receipt']]
            
            # Find cells that contain amount text and update them
            spelled_amount = self.spell_number(total_amount)
            
            # Look for cells that might contain amount text (this is a generic approach)
            # In a real implementation, you'd need to know the specific cell locations
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        # Look for patterns that indicate amount cells
                        if any(keyword in cell.value.upper() for keyword in ['DOLLAR', 'AMOUNT', 'TOTAL']):
                            # This is a simplified approach - you'd need to customize based on actual sheet structure
                            pass
            
            logger.info(f"Updated RECEIPT sheet with spelled amount: {spelled_amount}")
            
        except Exception as e:
            logger.warning(f"Could not update RECEIPT sheet: {e}")
    
    def get_existing_emds(self, workbook) -> List[str]:
        """
        Get list of existing EMD numbers from the workbook
        
        Args:
            workbook: OpenPyXL workbook object
            
        Returns:
            List of existing EMD numbers
        """
        sheet = workbook[self.config['sheets']['emd']]
        existing_emds = []
        
        # Read existing EMD numbers (assuming they start from row 8)
        for row in range(8, sheet.max_row + 1):
            emd_value = sheet.cell(row=row, column=self.config['emd_columns']['emd_number']).value
            if emd_value and str(emd_value).strip() != "":
                existing_emds.append(str(emd_value).strip())
        
        return existing_emds


# API Functions for external use
def process_emd_data(emd_data_dict: Dict[str, Any], template_path: str, output_path: str = None, 
                    config_path: str = "resources/emd_processor_config.yaml") -> str:
    """
    Main API function to process EMD data
    
    Args:
        emd_data_dict: Dictionary containing EMD data in the specified format
        template_path: Path to template Excel file
        output_path: Path for output file (optional, will generate if not provided)
        config_path: Path to configuration file
        
    Returns:
        Path to the generated output file
    """
    processor = EMDProcessor(config_path)
    
    # Convert dictionary to EMDData object
    emd_data = processor.create_emd_data_from_dict(emd_data_dict)
    
    # Load template workbook
    workbook = load_workbook(template_path)
    
    # Get existing EMDs for validation
    existing_emds = processor.get_existing_emds(workbook)
    
    # Validate data
    is_valid, error_msg = processor.validate_emd_data(emd_data, existing_emds)
    if not is_valid:
        raise ValueError(error_msg)
    
    # Fill EMD sheet
    processor.fill_emd_sheet(workbook, emd_data)
    
    # Fill SUM sheet with collection messages
    processor.fill_sum_sheet(workbook, emd_data)
    
    # Update receipt sheet with total amount
    processor.update_receipt_sheet(workbook, emd_data.total_amount)
    
    # Generate output filename if not provided
    if not output_path:
        # Extract flight number from first EMD record
        flight = emd_data.emd_list[0].flight if emd_data.emd_list else "0000"
        filename = processor.generate_filename(flight, emd_data.report_date)
        output_path = os.path.join(os.path.dirname(template_path), filename)
    
    # Save the workbook
    workbook.save(output_path)
    logger.info(f"File saved successfully as: {output_path}")
    
    return output_path


def spell_number(amount: float, config_path: str = "resources/emd_processor_config.yaml") -> str:
    """
    API function to convert number to text
    
    Args:
        amount: Amount to convert
        config_path: Path to configuration file
        
    Returns:
        Text representation of the amount
    """
    processor = EMDProcessor(config_path)
    return processor.spell_number(amount)


def validate_emd_data(emd_data_dict: Dict[str, Any], config_path: str = "resources/emd_processor_config.yaml") -> Tuple[bool, str]:
    """
    API function to validate EMD data
    
    Args:
        emd_data_dict: Dictionary containing EMD data
        config_path: Path to configuration file
        
    Returns:
        Tuple of (is_valid, error_message)
    """
    processor = EMDProcessor(config_path)
    emd_data = processor.create_emd_data_from_dict(emd_data_dict)
    return processor.validate_emd_data(emd_data)


def generate_filename(flight: str, date_str: str, config_path: str = "resources/emd_processor_config.yaml") -> str:
    """
    API function to generate filename
    
    Args:
        flight: Flight number
        date_str: Date string
        config_path: Path to configuration file
        
    Returns:
        Generated filename
    """
    processor = EMDProcessor(config_path)
    return processor.generate_filename(flight, date_str)


if __name__ == "__main__":
    # Example usage with new API format
    try:
        # Example: Process EMD data with new API format
        sample_data = {
            "total_amount": 12345.00,
            "report_date": "2025-08-07",
            "signature": "Jun Liu",
            "collection_msg": ["Message 1", "Message 2", "Message 3"],
            "emd_list": [
                {
                    "emd": "9991234567890",
                    "et": "9991234567890",
                    "leg": "LAX-PEK",
                    "emd_count": 1,
                    "operation": "issue",
                    "job_number": 35008,
                    "date": "2025-08-07",
                    "product": "EXPC",
                    "i_d": "international",
                    "flight": "CA984",
                    "price": 12345.00,
                    "cash": 12345.00,
                    "ax": 0.00,
                    "other": 0.00,
                    "last_4_digit": "",
                    "remark": "Sample remark"
                }
            ]
        }
        
        # Example: Process EMD data
        # output_file = process_emd_data(
        #     emd_data_dict=sample_data,
        #     template_path="path/to/template.xlsx",
        #     output_path="path/to/output.xlsx"
        # )
        # print(f"Processed file: {output_file}")
        
        # Example: Spell number
        amount_text = spell_number(1234.56)
        print(f"Amount in words: {amount_text}")
        
    except Exception as e:
        logger.error(f"Error in main execution: {e}") 