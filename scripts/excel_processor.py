#!/usr/bin/env python3
"""
Excel处理核心逻辑模块

约定：
- 输入Excel的表头在第2行（1-based），DataFrame应以该行为列名（使用 header=1 读取）。
- 列名与固定列序号严格对应，读取后进行校验。
- 本模块不依赖Streamlit，所有错误以异常或返回值方式传递给UI层。

用户习惯：中文注释；函数之间空两行；函数内不留空行。
"""

from __future__ import annotations

import os
import re
import sqlite3
from typing import Dict, List, Optional, Tuple
from datetime import date
import pandas as pd


# =============================
# 数据源列名与固定列序号定义（1-based）
# =============================
EXPECTED_COLUMNS: Dict[int, str] = {
    2: "EMD",
    3: "关联ET",
    4: "旅客姓名",
    5: "航班号",
    6: "航程",
    7: "航班日期",
    8: "报销单号",
    9: "退票单号",
    10: "操作",
    11: "实收金额",
    12: "币种",
    13: "代理费率",
    14: "代理费金额",
    15: "代理费币种",
    16: "付款方式",
    17: "交易流水号",
    18: "工作号",
    19: "操作时间",
    20: "产品类型",
    21: "产品详情",
    22: "票种",
}

# 全局航班信息（在处理期间设置，供其他模块直接引用）
FLIGHT_NUMBER: str = ''
FLIGHT_DATE: Optional[date] = None

def format_date_ddmmmyy(d: date) -> str:
    """将日期格式化为 DDMMMYY（英文大写月份缩写）"""
    MONTH_ABBR = ['JAN','FEB','MAR','APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC']
    return f"{d.day:02d}{MONTH_ABBR[d.month-1]}{d.year % 100:02d}"

def validate_input_columns(df_input: pd.DataFrame) -> None:
    """校验表头：第二行作为表头且与固定列序号完全一致。
    触发失败时抛出 ValueError 异常，消息包含第一个不匹配的位置。
    """
    columns = list(df_input.columns)
    for col_idx_1_based, expected_name in EXPECTED_COLUMNS.items():
        pos = col_idx_1_based - 1
        if pos >= len(columns):
            raise ValueError(f"列总数不足，缺少第{col_idx_1_based}列：应为“{expected_name}”")
        actual_name = str(columns[pos]).strip()
        if actual_name != expected_name:
            raise ValueError(
                f"第{col_idx_1_based}列列名不匹配：检测到“{actual_name}”，应为“{expected_name}”"
            )


def convert_to_string_no_decimal(value) -> str:
    """数字转字符串并去掉小数部分"""
    try:
        if not value or str(value).strip() == '' or str(value) == 'nan':
            return ""
        float_val = float(str(value))
        int_val = int(float_val)
        return str(int_val)
    except (ValueError, TypeError):
        return str(value) if value else ""


def extract_first_line(value) -> str:
    """提取多行文本的第一行"""
    try:
        if not value or str(value).strip() == '' or str(value) == 'nan':
            return ""
        value_str = str(value).strip()
        first_line = value_str.split('\n')[0].strip()
        return first_line
    except Exception:
        return str(value) if value else ""


def translate_column_t_to_h(value: str) -> str:
    """产品类型到H列的翻译"""
    if not value or str(value).strip() == '' or str(value) == 'nan':
        return ""
    value_str = str(value).strip()
    translation_map = {
        '逾重PC': 'EXPC',
        '选座': 'SEAT',
        '升舱': 'UPG',
    }
    return translation_map.get(value_str, value_str)


def translate_operation_to_english(value: str) -> str:
    """操作列中文到英文翻译"""
    if not value or str(value).strip() == '' or str(value) == 'nan':
        return ""
    value_str = str(value).strip()
    operation_translation = {
        '出票': 'Issue',
        '废票': 'Void',
    }
    return operation_translation.get(value_str, value_str)


def has_ckin_ccrd(record: Dict) -> bool:
    """是否包含CKIN CCRD信息"""
    ckin_msg = record.get('ckin_msg', '')
    return 'CKIN CCRD' in ckin_msg


def parse_ckin_ccrd(ckin_msg: str) -> Dict:
    """解析CKIN CCRD内容，返回是否成功及L/M/N/O/P列数据"""
    try:
        if not ckin_msg or 'CKIN CCRD' not in ckin_msg:
            return {'success': False, 'data': {}}
        ckin_pattern = r'CKIN CCRD\s+([^;]+)'
        matches = re.findall(ckin_pattern, ckin_msg, re.IGNORECASE)
        if not matches:
            return {'success': False, 'data': {}}
        data = {'L': '', 'M': '', 'N': '', 'O': '', 'P': ''}
        for ccrd_content in matches:
            ccrd_content = ccrd_content.strip()
            parts = ccrd_content.split()
            if len(parts) < 1:
                continue
            item1 = parts[0]
            item2 = parts[1] if len(parts) > 1 else ''
            item3_and_beyond = ' '.join(parts[2:]) if len(parts) > 2 else ''
            if item1.upper() == "CASH":
                data['L'] = item2
                return {'success': True, 'data': data}
            elif re.match(r'^[A-Z]{2}\d{4}$', item1):
                letters = item1[:2]
                digits = item1[2:]
                data['O'] = digits
                if letters.upper() == 'AX':
                    data['M'] = item2
                else:
                    data['N'] = item2
                data['P'] = item3_and_beyond
                return {'success': True, 'data': data}
        return {'success': False, 'data': {}}
    except Exception:
        return {'success': False, 'data': {}}


def extract_ckin_ccrd_content(ckin_msg: str) -> str:
    """提取CKIN CCRD后面的内容直到分号"""
    try:
        if not ckin_msg or 'CKIN CCRD' not in ckin_msg:
            return ckin_msg
        pattern = r'CKIN CCRD\s+([^;]+)'
        match = re.search(pattern, ckin_msg)
        if match:
            return match.group(1).strip()
        else:
            return ckin_msg
    except Exception:
        return ckin_msg


def number_to_english(amount: float) -> str:
    """数字金额转英文表达"""
    try:
        ones = ['', 'ONE', 'TWO', 'THREE', 'FOUR', 'FIVE', 'SIX', 'SEVEN', 'EIGHT', 'NINE']
        teens = ['TEN', 'ELEVEN', 'TWELVE', 'THIRTEEN', 'FOURTEEN', 'FIFTEEN', 'SIXTEEN', 'SEVENTEEN', 'EIGHTEEN', 'NINETEEN']
        tens = ['', '', 'TWENTY', 'THIRTY', 'FORTY', 'FIFTY', 'SIXTY', 'SEVENTY', 'EIGHTY', 'NINETY']
        def convert_hundreds(n):
            result = ''
            if n >= 100:
                result += ones[n // 100] + ' HUNDRED '
                n %= 100
            if n >= 20:
                result += tens[n // 10] + ' '
                n %= 10
            elif n >= 10:
                result += teens[n - 10] + ' '
                n = 0
            if n > 0:
                result += ones[n] + ' '
            return result.strip()
        dollars = int(amount)
        cents = int(round((amount - dollars) * 100))
        result = ''
        if dollars == 0:
            result = 'ZERO DOLLARS'
        else:
            if dollars >= 1000:
                thousands = dollars // 1000
                result += convert_hundreds(thousands) + ' THOUSAND '
                dollars %= 1000
            if dollars > 0:
                result += convert_hundreds(dollars)
            if int(amount) == 1:
                result += ' DOLLAR'
            else:
                result += ' DOLLARS'
        if cents > 0:
            if cents == 1:
                result += ' AND ' + convert_hundreds(cents) + ' CENT'
            else:
                result += ' AND ' + convert_hundreds(cents) + ' CENTS'
        else:
            result += ' AND NO CENTS'
        result += ' EXACTLY'
        return result.strip()
    except Exception:
        return f"${amount:.2f}"


def get_all_ckin_ccrd_hbnb(db) -> List[Dict]:
    """查询所有包含CKIN CCRD的HBNB记录"""
    conn = sqlite3.connect(db.db_file)
    cursor = conn.cursor()
    query = (
        """
        SELECT hbnb_number, name, tkne, ckin_msg 
        FROM hbpr_full_records 
        WHERE ckin_msg IS NOT NULL 
        AND ckin_msg LIKE '%CKIN CCRD%'
        AND ckin_msg != ''
        """
    )
    cursor.execute(query)
    records = cursor.fetchall()
    conn.close()
    return [
        {
            'hbnb_number': record[0],
            'name': record[1] or '未知',
            'tkne': record[2] or '',
            'ckin_msg': record[3] or ''
        }
        for record in records
    ]


def find_records_by_tkne(db, tkne: str) -> List[Dict]:
    """根据TKNE查找数据库记录"""
    conn = sqlite3.connect(db.db_file)
    cursor = conn.cursor()
    clean_tkne = str(tkne).replace('.0', '') if tkne else ''
    query = (
        """
        SELECT hbnb_number, name, tkne, ckin_msg 
        FROM hbpr_full_records 
        WHERE tkne LIKE ? OR tkne LIKE ? OR tkne = ?
        """
    )
    patterns = [f'{clean_tkne}/1', f'{clean_tkne}/2', clean_tkne]
    cursor.execute(query, patterns)
    records = cursor.fetchall()
    conn.close()
    return [
        {
            'hbnb_number': record[0],
            'name': record[1] or '未知',
            'tkne': record[2] or '',
            'ckin_msg': record[3] or ''
        }
        for record in records
    ]


def create_base_output_row(input_row: pd.Series) -> Dict:
    """创建输出基础行（映射到A-P列），第7列(G)使用“航班日期”。"""
    output_row: Dict[str, str] = {}
    emd_value = input_row.get('EMD', '')
    output_row['A'] = convert_to_string_no_decimal(emd_value)
    et_value = input_row.get('关联ET', '')
    output_row['B'] = convert_to_string_no_decimal(et_value)
    route_value = input_row.get('航程', '')
    output_row['C'] = extract_first_line(route_value)
    operation_value = str(input_row.get('操作', ''))
    output_row['E'] = translate_operation_to_english(operation_value)
    job_no_value = input_row.get('工作号', '')
    output_row['F'] = convert_to_string_no_decimal(job_no_value)
    # 设置全局航班信息（仅在首次赋值时）
    global FLIGHT_DATE, FLIGHT_NUMBER
    if FLIGHT_DATE is None or FLIGHT_DATE == '':
        try:
            parsed_date = pd.to_datetime(input_row.get('航班日期', ''), errors='coerce')
            if parsed_date is not None and not pd.isna(parsed_date):
                FLIGHT_DATE = parsed_date.date()
        except Exception:
            FLIGHT_DATE = None
    if not FLIGHT_NUMBER:
        FLIGHT_NUMBER = str(input_row.get('航班号', '') or '').strip()
    # 输出列赋值
    output_row['G'] = str(input_row.get('航班日期', ''))
    output_row['J'] = extract_first_line(input_row.get('航班号', ''))
    output_row['K'] = str(input_row.get('实收金额', ''))
    product_type = str(input_row.get('产品类型', ''))
    output_row['H'] = translate_column_t_to_h(product_type)
    output_row['D'] = "1"
    output_row['I'] = "International"
    output_row['L'] = ""
    output_row['M'] = ""
    output_row['N'] = ""
    output_row['O'] = ""
    output_row['P'] = ""
    return output_row


def process_excel_file(df_input: pd.DataFrame, db, debug: bool = False) -> Tuple[Optional[pd.DataFrame], List[Dict], List[Dict]]:
    """处理输入数据，返回结果表、未处理记录、调试日志（按行）。

    注：全局变量 FLIGHT_NUMBER/FLIGHT_DATE 会在处理期间被设置，可供其他地方直接使用。
    """
    # 校验表头
    validate_input_columns(df_input)
    output_data: List[Dict] = []
    unprocessed_records: List[Dict] = []
    debug_logs: List[Dict] = []
    hbnb_list = get_all_ckin_ccrd_hbnb(db)
    void_emds: set[str] = set()
    for _, row in df_input.iterrows():
        operation = str(row.get('操作', '')).strip()
        if operation in ['废票', 'Void']:
            emd_number = str(row.get('EMD', '')).strip()
            if emd_number and emd_number != 'nan':
                void_emd_clean = convert_to_string_no_decimal(emd_number)
                void_emds.add(void_emd_clean)
    for index, row in df_input.iterrows():
        try:
            tkne = str(row.get('关联ET', '')).strip()
            if not tkne or tkne == 'nan':
                continue
            hbnb_records = find_records_by_tkne(db, tkne)
            output_row = create_base_output_row(row)
            current_emd = convert_to_string_no_decimal(str(row.get('EMD', '')))
            if current_emd in void_emds:
                output_row['D'] = 0
                output_row['K'] = 0
            else:
                for hbnb_record in hbnb_records:
                    if has_ckin_ccrd(hbnb_record):
                        ckin_data = parse_ckin_ccrd(hbnb_record['ckin_msg'])
                        if ckin_data['success']:
                            output_row.update(ckin_data['data'])
                            hbnb_list = [h for h in hbnb_list if h['hbnb_number'] != hbnb_record['hbnb_number']]
                            break
                        else:
                            unprocessed_records.append({
                                'name': hbnb_record.get('name', '未知'),
                                'tkne': tkne,
                                'ckin_ccrd': hbnb_record['ckin_msg']
                            })
            output_data.append(output_row)
            if debug:
                debug_logs.append({
                    'row_index': index + 1,
                    'input': {k: row.get(k, '') for k in df_input.columns},
                    'output': output_row.copy(),
                })
        except Exception as e:
            unprocessed_records.append({'name': '未知', 'tkne': tkne if 'tkne' in locals() else '', 'ckin_ccrd': f'行{index+1}处理错误: {str(e)}'})
            if debug:
                debug_logs.append({'row_index': index + 1, 'error': str(e)})
            continue
    for remaining_hbnb in hbnb_list:
        unprocessed_records.append({
            'name': remaining_hbnb.get('name', '未知'),
            'tkne': remaining_hbnb.get('tkne', '未知'),
            'ckin_ccrd': remaining_hbnb['ckin_msg']
        })
    if output_data:
        result_df = pd.DataFrame(output_data)
        return result_df, unprocessed_records, debug_logs
    return None, unprocessed_records, debug_logs


def generate_output_excel(result_df: pd.DataFrame, unprocessed_records: List[Dict], output_file: str) -> str:
    """根据模板生成输出Excel文件，返回保存路径。

    使用全局 FLIGHT_NUMBER/FLIGHT_DATE 写入 SUM 表。
    """
    from openpyxl import load_workbook
    template_file = os.path.join("resources", "Out_format.xlsx")
    if not os.path.exists(template_file):
        raise FileNotFoundError(f"模板文件不存在: {template_file}")
    wb = load_workbook(template_file)
    if 'EMD' not in wb.sheetnames:
        raise ValueError("模板文件中没有EMD工作表")
    ws_emd = wb['EMD']
    start_row = 8
    headers = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']
    for data_idx, (_, row) in enumerate(result_df.iterrows()):
        row_idx = start_row + data_idx
        for col_idx, header in enumerate(headers, 1):
            value = row.get(header, '')
            if col_idx == 4:
                if str(value) == "0":
                    ws_emd.cell(row=row_idx, column=col_idx, value=0)
                else:
                    ws_emd.cell(row=row_idx, column=col_idx, value=1)
            elif col_idx in [11, 12, 13, 14]:
                try:
                    if value and str(value).strip() and str(value) != 'nan':
                        numeric_value = float(str(value).strip())
                    else:
                        numeric_value = 0
                    ws_emd.cell(row=row_idx, column=col_idx, value=numeric_value)
                except (ValueError, TypeError):
                    ws_emd.cell(row=row_idx, column=col_idx, value=0)
            elif col_idx == 15:
                if value and str(value).strip() and str(value) != 'nan':
                    ws_emd.cell(row=row_idx, column=col_idx, value=str(value).strip())
                else:
                    ws_emd.cell(row=row_idx, column=col_idx, value='')
            else:
                ws_emd.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else '')
    if 'SUM' not in wb.sheetnames:
        ws_sum = wb.create_sheet('SUM')
    else:
        ws_sum = wb['SUM']
    # 写入 K4：航班号数字部分（优先使用全局 FLIGHT_NUMBER）
    fn_for_digits = None
    if FLIGHT_NUMBER:
        fn_for_digits = FLIGHT_NUMBER
    elif len(result_df) > 0:
        candidate = result_df.iloc[0].get('J', '')
        if candidate and isinstance(candidate, str):
            fn_for_digits = candidate
    if fn_for_digits:
        flight_digits = re.findall(r'\d+', fn_for_digits)
        if flight_digits:
            ws_sum.cell(row=4, column=11, value=flight_digits[0])
    # 将SUM中C14写为航班日期（来自结果表第7列G）
    if FLIGHT_DATE is not None:
        ws_sum.cell(row=14, column=3, value=FLIGHT_DATE.strftime('%Y-%m-%d'))
    else:
        # 回退：从结果数据G列尝试
        flight_date_cell_value = None
        if 'G' in result_df.columns and len(result_df) > 0:
            for _, r in result_df.iterrows():
                candidate = r.get('G', '')
                if candidate is not None and str(candidate).strip() and str(candidate).strip() != 'nan':
                    flight_date_cell_value = str(candidate).strip()
                    break
        ws_sum.cell(row=14, column=3, value=flight_date_cell_value or '')
    if unprocessed_records:
        row_idx = 15
        for record in unprocessed_records:
            ckin_ccrd_content = extract_ckin_ccrd_content(record.get('ckin_ccrd', ''))
            error_info = f"乘客: {record.get('name', '未知')}, TKNE: {record.get('tkne', '')}, {ckin_ccrd_content}"
            ws_sum.cell(row=row_idx, column=3, value=error_info)
            row_idx += 1
    if 'RECEIPT' in wb.sheetnames:
        ws_receipt = wb['RECEIPT']
        cash_total = 0.0
        for _, row in result_df.iterrows():
            l_value = row.get('L', '')
            if l_value and str(l_value).strip() and str(l_value).strip() != 'nan':
                try:
                    cash_amount = float(str(l_value).strip())
                    cash_total += cash_amount
                except ValueError:
                    continue
        if cash_total > 0:
            english_amount = number_to_english(cash_total)
            ws_receipt.cell(row=8, column=3, value=english_amount)
        if len(result_df) > 0:
            route = result_df.iloc[0].get('C', '')
            flight = result_df.iloc[0].get('J', '')
            if route and '-' in route:
                departure = route.split('-')[0][:3].upper()
                ws_receipt.cell(row=17, column=8, value=departure)
                if flight:
                    report_title = f"{departure} {flight} EMD REPORT"
                    ws_receipt.cell(row=14, column=3, value=report_title)
    wb.save(output_file)
    wb.close()
    return output_file


